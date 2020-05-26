import openpyxl

wb = openpyxl.load_workbook('Daily.xlsx')

print(wb.sheetnames)

sheet1 = wb["Sheet1"]

buyStartr = 3
buyStartc = 3
buyEndr = 4
buyEndc = 4

sellStartr = 3
sellStartc = 10
sellEndr = 4
sellEndc = 11

def updateBuy():
    for i in range(buyStartr, buyEndr + 1):
        qty = sheet1.cell(row=i, column=buyStartc).value
        cost = sheet1.cell(row=i, column=buyEndc).value
        sheet1.cell(row=i, column=buyEndc + 1).value = qty * cost

def updateSell():
    for i in range(sellStartr, sellEndr + 1):
        qty = sheet1.cell(row=i, column=sellStartc).value
        cost = sheet1.cell(row=i, column=sellEndc).value
        sheet1.cell(row=i, column=sellEndc + 1).value = qty * cost


updateBuy()
updateSell()
wb.save("New.xlsx")